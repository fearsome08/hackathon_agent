import argparse
import json
import os
import re
import tkinter as tk
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, ttk
from typing import Iterable
from urllib.parse import urljoin, urlparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright


@dataclass
class PageParameter:
    page_url: str
    kind: str
    name: str
    label: str
    input_type: str
    value: str
    options: list[str]


@dataclass
class ExtractedField:
    page_url: str
    name: str
    value: str
    source: str


@dataclass
class WorkItemLink:
    page_url: str
    text: str
    href: str
    context: str
    section: str = ""


@dataclass
class ImplementsRequirementItem:
    link_text: str
    link_href: str
    fields: list[ExtractedField] = field(default_factory=list)


@dataclass
class LinkedWorkItem:
    source_link_text: str
    source_link_href: str
    fields: list[ExtractedField] = field(default_factory=list)
    implements_requirements: list[ImplementsRequirementItem] = field(default_factory=list)


@dataclass
class AgentResult:
    fields: list[ExtractedField] = field(default_factory=list)
    parameters: list[PageParameter] = field(default_factory=list)
    links: list[WorkItemLink] = field(default_factory=list)
    swe1_items: list[LinkedWorkItem] = field(default_factory=list)


@dataclass
class AgentConfig:
    login_url: str
    start_url: str | None
    username: str
    password: str
    pages: int
    show_browser: bool
    manual_login: bool
    report_path: str
    browser: str
    swe1_keyword: str = ""


def env(name: str, default: str | None = None) -> str:
    value = os.getenv(name, default)
    if value is None or value == "":
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def prompt_for_config() -> AgentConfig:
    result: dict[str, object] = {}

    root = tk.Tk()
    root.title("Corporate Website Agent")
    root.resizable(False, False)

    frame = ttk.Frame(root, padding=16)
    frame.grid(row=0, column=0, sticky="nsew")

    login_url_var = tk.StringVar()
    start_url_var = tk.StringVar()
    username_var = tk.StringVar(value=os.getenv("CORP_USERNAME", ""))
    password_var = tk.StringVar(value=os.getenv("CORP_PASSWORD", ""))
    pages_var = tk.StringVar(value="1")
    report_path_var = tk.StringVar(value="work_item_report.md")
    browser_var = tk.StringVar(value="msedge")
    show_browser_var = tk.BooleanVar(value=True)
    manual_login_var = tk.BooleanVar(value=False)

    fields = [
        ("Login URL", login_url_var, False),
        ("Page to inspect", start_url_var, False),
        ("Username", username_var, False),
        ("Password", password_var, True),
        ("Pages", pages_var, False),
        ("Report file", report_path_var, False),
    ]

    for row, (label, variable, secret) in enumerate(fields):
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", pady=4)
        entry = ttk.Entry(frame, width=54, textvariable=variable, show="*" if secret else "")
        entry.grid(row=row, column=1, sticky="ew", pady=4)

    browser_row = len(fields)
    ttk.Label(frame, text="Browser").grid(row=browser_row, column=0, sticky="w", pady=4)
    browser_select = ttk.Combobox(
        frame,
        width=51,
        textvariable=browser_var,
        values=["msedge", "chromium"],
        state="readonly",
    )
    browser_select.grid(row=browser_row, column=1, sticky="ew", pady=4)

    ttk.Checkbutton(frame, text="Show browser while running", variable=show_browser_var).grid(
        row=len(fields) + 1, column=1, sticky="w", pady=(8, 4)
    )
    ttk.Checkbutton(frame, text="Manual login / SSO", variable=manual_login_var).grid(
        row=len(fields) + 2, column=1, sticky="w", pady=4
    )

    note = ttk.Label(frame, text="Leave Page to inspect empty to inspect the page reached after login.")
    note.grid(row=len(fields) + 3, column=0, columnspan=2, sticky="w", pady=(4, 10))

    def submit() -> None:
        try:
            pages = int(pages_var.get())
            if pages < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid pages", "Pages must be a number greater than 0.")
            return

        login_url = login_url_var.get().strip()
        username = username_var.get().strip()
        password = password_var.get()
        manual_login = manual_login_var.get()

        if not login_url:
            messagebox.showerror("Missing details", "Login URL is required.")
            return

        if not manual_login and (not username or not password):
            messagebox.showerror("Missing details", "Username and password are required unless Manual login / SSO is checked.")
            return

        result["config"] = AgentConfig(
            login_url=login_url,
            start_url=start_url_var.get().strip() or None,
            username=username,
            password=password,
            pages=pages,
            show_browser=show_browser_var.get() or manual_login,
            manual_login=manual_login,
            report_path=report_path_var.get().strip() or "work_item_report.md",
            browser=browser_var.get(),
            swe1_keyword=os.getenv("SWE1_KEYWORD", ""),
        )
        root.destroy()

    def cancel() -> None:
        root.destroy()

    buttons = ttk.Frame(frame)
    buttons.grid(row=len(fields) + 4, column=0, columnspan=2, sticky="e")
    ttk.Button(buttons, text="Cancel", command=cancel).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(buttons, text="Run Agent", command=submit).grid(row=0, column=1)

    root.bind("<Return>", lambda _event: submit())
    root.bind("<Escape>", lambda _event: cancel())
    root.mainloop()

    config = result.get("config")
    if not isinstance(config, AgentConfig):
        raise SystemExit("Cancelled.")
    return config


def same_domain(url: str, root_url: str) -> bool:
    return urlparse(url).netloc == urlparse(root_url).netloc


def clean_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


WORK_ITEM_FIELD_LABELS = [
    "Type",
    "Filed Against",
    "Priority",
    "Owned By",
    "Planned For",
    "Target Configuration",
    "Tags",
    "Description",
    "Due Date",
    "Creation Date",
]

# Fields expected in the DNG MPCI_standard view
DNG_FIELD_LABELS = [
    "Artifact Type",
    "Feature",
    "SW_Function",
    "Status",
    "ValidSince",
    "InvalidSince",
    "Comment_Internal",
    "Safety",
    "Security",
    "Legal",
    "Verification_Level",
    "Verification_Criteria",
    "Satisfies",
    "Link To",
    "Satisfied by Architecture Element",
    "Implemented By",
    "Validated By",
    "RT0ID",
    "Change Request Link",
    "Resource",
]


def normalize_label(value: str) -> str:
    return clean_text(value).rstrip(":").rstrip("*").strip()


def extract_work_item_id(url: str) -> str | None:
    match = re.search(r"(?:#|&)id=([^&#]+)", url)
    if match:
        return match.group(1)
    return None


def build_work_item_tab_urls(url: str) -> list[str]:
    if "action=com.ibm.team.workitem.viewWorkItem" not in url:
        return []

    work_item_id = extract_work_item_id(url)
    if not work_item_id:
        return []

    base_url = url.split("#", 1)[0] + f"#action=com.ibm.team.workitem.viewWorkItem&id={work_item_id}"
    return [
        base_url,
        base_url + "&tab=com.ibm.team.workitem.tab.history",
        base_url + "&tab=com.ibm.team.workitem.tab.links",
    ]


def extract_known_fields_from_text(page: Page, labels: list[str]) -> list[ExtractedField]:
    try:
        body_text = page.locator("body").inner_text(timeout=5_000)
    except Exception:
        return []

    lines = [clean_text(line) for line in body_text.splitlines()]
    lines = [line for line in lines if line]
    normalized_to_label = {normalize_label(label).lower(): label for label in labels}

    extracted: list[ExtractedField] = []
    seen: set[str] = set()

    index = 0
    while index < len(lines):
        line = lines[index]
        normalized_line = normalize_label(line).lower()

        label = normalized_to_label.get(normalized_line)
        inline_value = ""
        if label is None:
            for known_normalized, known_label in normalized_to_label.items():
                prefix_patterns = [
                    f"{known_normalized}:",
                    f"{known_normalized}: *",
                    f"{known_normalized} *:",
                ]
                lower_line = line.lower()
                for prefix in prefix_patterns:
                    if lower_line.startswith(prefix):
                        label = known_label
                        inline_value = clean_text(line[len(prefix) :])
                        break
                if label:
                    break

        if label is None:
            index += 1
            continue

        value_parts: list[str] = []
        if inline_value:
            value_parts.append(inline_value)

        next_index = index + 1
        max_value_lines = 10  # cap to avoid consuming sidebar/module content
        lines_consumed = 0
        while next_index < len(lines) and lines_consumed < max_value_lines:
            candidate = lines[next_index]
            if normalize_label(candidate).lower() in normalized_to_label:
                break
            if any(candidate.lower().startswith(f"{known}:") for known in normalized_to_label):
                break
            value_parts.append(candidate)
            next_index += 1
            lines_consumed += 1

        value = clean_text(" ".join(value_parts))
        if label not in seen:
            extracted.append(ExtractedField(page_url=page.url, name=label, value=value, source="visible_text"))
            seen.add(label)

        index = max(next_index, index + 1)

    return extracted


def fields_from_parameters(parameters: list[PageParameter]) -> list[ExtractedField]:
    fields_by_name: dict[str, ExtractedField] = {}
    known_labels = {normalize_label(label).lower(): label for label in WORK_ITEM_FIELD_LABELS}

    for parameter in parameters:
        candidates = [parameter.label, parameter.name]
        matched_label = None
        for candidate in candidates:
            normalized = normalize_label(candidate).lower()
            if normalized in known_labels:
                matched_label = known_labels[normalized]
                break

        if matched_label is None:
            continue

        value = parameter.value
        if not value and parameter.options:
            value = ", ".join(parameter.options)

        if matched_label not in fields_by_name:
            fields_by_name[matched_label] = ExtractedField(
                page_url=parameter.page_url,
                name=matched_label,
                value=value,
                source="form_control",
            )

    return list(fields_by_name.values())


def merge_fields(fields: list[ExtractedField]) -> list[ExtractedField]:
    merged: dict[tuple[str, str], ExtractedField] = {}
    for field_value in fields:
        key = (field_value.page_url, field_value.name)
        existing = merged.get(key)
        if existing is None or (not existing.value and field_value.value):
            merged[key] = field_value
    return list(merged.values())


def first_visible(page: Page, selectors: Iterable[str]):
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            if locator.is_visible(timeout=1_000):
                return locator
        except PlaywrightTimeoutError:
            continue
    return None


def login(page: Page, login_url: str, username: str, password: str) -> None:
    page.goto(login_url, wait_until="domcontentloaded")

    username_input = first_visible(
        page,
        [
            "input[name='username']",
            "input[name='email']",
            "input[type='email']",
            "input[id*='user' i]",
            "input[name*='user' i]",
            "input[id*='email' i]",
            "input[name*='email' i]",
        ],
    )
    password_input = first_visible(
        page,
        [
            "input[name='password']",
            "input[type='password']",
            "input[id*='pass' i]",
            "input[name*='pass' i]",
        ],
    )

    if username_input is None or password_input is None:
        raise RuntimeError("Could not find visible username/password fields. Add site-specific selectors.")

    username_input.fill(username)
    password_input.fill(password)

    submit = first_visible(
        page,
        [
            "button[type='submit']",
            "input[type='submit']",
            "button:has-text('Log in')",
            "button:has-text('Login')",
            "button:has-text('Sign in')",
            "button:has-text('Continue')",
        ],
    )
    if submit is None:
        password_input.press("Enter")
    else:
        submit.click()

    page.wait_for_load_state("networkidle", timeout=30_000)


def wait_for_manual_login(page: Page, login_url: str) -> None:
    page.goto(login_url, wait_until="domcontentloaded")
    print("")
    print("Manual login mode:")
    print("1. Complete the login in the browser window.")
    print("2. Navigate to the page you want to inspect if needed.")
    print("3. Return here and press Enter.")
    input("Press Enter after login is complete...")
    page.wait_for_load_state("networkidle", timeout=30_000)


def label_for_element(element) -> str:
    return element.evaluate(
        """
        el => {
          const labels = el.labels ? Array.from(el.labels).map(l => l.innerText.trim()) : [];
          if (labels.length) return labels.join(" ");
          const aria = el.getAttribute("aria-label");
          if (aria) return aria;
          const labelledBy = el.getAttribute("aria-labelledby");
          if (labelledBy) {
            return labelledBy.split(/\\s+/)
              .map(id => document.getElementById(id)?.innerText?.trim())
              .filter(Boolean)
              .join(" ");
          }
          const placeholder = el.getAttribute("placeholder");
          if (placeholder) return placeholder;
          const parentText = el.closest("label")?.innerText?.trim();
          if (parentText) return parentText;
          return "";
        }
        """,
    )


def extract_parameters(page: Page) -> list[PageParameter]:
    # Batch-extract all element data in a single JS round-trip for speed
    raw: list[dict] = page.evaluate(
        """
        () => {
          const results = [];
          const selector = 'input, select, textarea, button, [role="button"], [role="combobox"]';
          const elements = Array.from(document.querySelectorAll(selector));
          elements.forEach((el, index) => {
            try {
              const rect = el.getBoundingClientRect();
              if (rect.width === 0 && rect.height === 0) return;
              const style = window.getComputedStyle(el);
              if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0') return;

              const tag = el.tagName.toLowerCase();
              const role = el.getAttribute('role') || '';
              const kind = role || tag;
              const inputType = el.getAttribute('type') || tag;
              const name = el.getAttribute('name') || el.getAttribute('id') || el.getAttribute('data-testid') || (kind + '_' + (index + 1));

              let value = '';
              if (tag === 'input' || tag === 'textarea') value = el.value || '';
              else if (tag === 'select') value = el.value || '';
              else value = (el.innerText || '').trim().replace(/\\s+/g, ' ');

              let label = (el.getAttribute('aria-label') || el.getAttribute('placeholder') || '').trim();
              if (!label) {
                const labels = el.labels ? Array.from(el.labels).map(l => l.innerText.trim()) : [];
                if (labels.length) label = labels.join(' ');
              }
              if (!label) {
                const lbId = el.getAttribute('aria-labelledby');
                if (lbId) label = lbId.split(/\\s+/).map(id => document.getElementById(id)?.innerText?.trim()).filter(Boolean).join(' ');
              }
              if (!label) label = el.closest('label')?.innerText?.trim() || '';

              const options = tag === 'select'
                ? Array.from(el.options).map(o => o.text.trim()).filter(Boolean)
                : [];

              results.push({ kind, name, label, inputType, value, options });
            } catch (e) {}
          });
          return results;
        }
        """
    )

    page_url = page.url
    return [
        PageParameter(
            page_url=page_url,
            kind=clean_text(r["kind"]),
            name=clean_text(r["name"]),
            label=clean_text(r["label"]),
            input_type=clean_text(r["inputType"]),
            value=clean_text(r["value"]),
            options=[clean_text(o) for o in r["options"] if o.strip()],
        )
        for r in raw
    ]


def extract_links_tab_entries(page: Page) -> list[WorkItemLink]:
    """Extract all links from the current page (must already be on the Links Tab)."""
    page_url = page.url

    raw: list[dict] = page.evaluate(
        """
        () => {
          const results = [];
          const seen = new Set();

          const isVisible = el => {
            const r = el.getBoundingClientRect();
            if (r.width === 0 && r.height === 0) return false;
            const s = window.getComputedStyle(el);
            return s.display !== 'none' && s.visibility !== 'hidden';
          };

          const SECTION_RE = /implement|depend|track|block|relat|child|parent|duplicate|resolv|reference|contribut|affect|test/i;
          const headingTags = new Set(['h1','h2','h3','h4','h5','h6']);
          const headingRoles = new Set(['heading','columnheader','rowheader','group','separator']);

          let currentSection = '';

          const walker = document.createTreeWalker(
            document.body,
            NodeFilter.SHOW_ELEMENT,
            null
          );

          let node = walker.nextNode();
          while (node) {
            const tag = node.tagName.toLowerCase();
            const role = (node.getAttribute('role') || '').toLowerCase();

            // Always update section on explicit heading elements — this ensures
            // section resets when moving past 'Implements Requirement' to next group.
            if (headingTags.has(tag) || headingRoles.has(role)) {
              const txt = (node.innerText || node.textContent || '').trim().replace(/\\s+/g,' ');
              if (txt && txt.length < 150) currentSection = txt;
            }

            // EWM section label cells: short own-text, no child anchor links
            if (['td','th','div','span','li'].includes(tag) && !node.querySelector('a[href]')) {
              const ownText = Array.from(node.childNodes)
                .filter(n => n.nodeType === 3)
                .map(n => n.textContent.trim())
                .join(' ').trim();
              if (ownText && ownText.length < 80) {
                currentSection = ownText;
              } else {
                const fullText = (node.innerText || '').trim().replace(/\\s+/g,' ');
                if (fullText && fullText.length < 60) currentSection = fullText;
              }
            }

            if (tag === 'a' && node.getAttribute('href')) {
              const href = node.getAttribute('href');
              if (href && !href.startsWith('javascript:') && isVisible(node)) {
                const text = (node.innerText || node.textContent || '').trim().replace(/\\s+/g,' ');
                const container = node.closest('[role="row"], tr, li, section, article') || node.parentElement;
                const context = (container?.innerText || '').trim().replace(/\\s+/g,' ');
                const key = text + '|' + href + '|' + currentSection;
                if (!seen.has(key)) {
                  seen.add(key);
                  results.push({ text, href, context, section: currentSection });
                }
              }
            }

            node = walker.nextNode();
          }
          return results;
        }
        """
    )

    extracted: list[WorkItemLink] = []
    for r in raw:
        href = r.get("href", "")
        if not href:
            continue
        extracted.append(
            WorkItemLink(
                page_url=page_url,
                text=r.get("text", ""),
                href=urljoin(page_url, href),
                context=r.get("context", ""),
                section=r.get("section", ""),
            )
        )
    return extracted


def is_swe1_link(link: WorkItemLink, keyword: str = "") -> bool:
    """Return True only for genuine SWE1 work item links.

    Requires:
    - '[SWE' with a bracket in the link text (avoids nav tabs and user links)
    - href points to a CCM work item (viewWorkItem or itemOid/WorkItem)
    """
    text_lower = link.text.lower()
    href_lower = link.href.lower()

    # Must look like a CCM work item URL
    is_work_item_href = (
        "viewworkitem" in href_lower
        or "itemoid/com.ibm.team.workitem" in href_lower
    )
    if not is_work_item_href:
        return False

    # Must have [SWE (bracketed) in the link text itself
    if "[swe" not in text_lower:
        return False

    # Optional keyword filter
    if keyword:
        combined = f"{link.text} {link.context} {link.section}".lower()
        return keyword.lower() in combined

    return True


def is_implements_requirement_link(link: WorkItemLink) -> bool:
    """Return True only if the section exactly matches 'Implements Requirement'
    AND the href is a DNG resource URL."""
    # Strip trailing count like " (35)" from section label
    section = re.sub(r'\s*\(\d+\)\s*$', '', link.section).strip().lower()
    if section != "implements requirement":
        return False
    # Must be a DNG artifact URL — eliminates CCM, user profiles, Confluence, etc.
    return "/rm/resources/" in link.href or "/rm/views/" in link.href


def resolve_view_url(page: Page, href: str) -> str:
    """Navigate to href and return the final URL after any redirects."""
    page.goto(href, wait_until="load", timeout=30_000)
    page.wait_for_timeout(800)  # allow SPA to settle
    return page.url


def goto_and_settle(page: Page, url: str, timeout: int = 30_000) -> None:
    """Navigate and wait for the page to settle (faster than networkidle)."""
    page.goto(url, wait_until="load", timeout=timeout)
    page.wait_for_timeout(1200)


def build_linked_item_tab_urls(resolved_url: str) -> list[str]:
    """Like build_work_item_tab_urls but skips the history tab for speed."""
    if "action=com.ibm.team.workitem.viewWorkItem" not in resolved_url:
        return []
    work_item_id = extract_work_item_id(resolved_url)
    if not work_item_id:
        return []
    base_url = resolved_url.split("#", 1)[0] + f"#action=com.ibm.team.workitem.viewWorkItem&id={work_item_id}"
    return [
        base_url,
        base_url + "&tab=com.ibm.team.workitem.tab.links",
    ]


def apply_dng_view(url: str, view_name: str = "MPCI_standard") -> str:
    """Return the URL unchanged — view selection is done via the DNG UI after navigation."""
    return url


def select_dng_view(page: Page, view_name: str = "MPCI_standard") -> None:
    """Try to select a saved view by name in DNG's view dropdown."""
    try:
        # DNG renders a view switcher button/dropdown — try common selectors
        view_selectors = [
            f"[title*='{view_name}']",
            f"[aria-label*='{view_name}']",
            f"option[value*='{view_name}']",
            f"li:has-text('{view_name}')",
            f"span:has-text('{view_name}')",
        ]
        # First try to open the view menu
        menu_selectors = [
            "[title*='View']",
            "[aria-label*='view']",
            "button:has-text('View')",
            "[class*='viewMenu']",
            "[class*='view-menu']",
            "[class*='ViewSelector']",
        ]
        menu_opened = False
        for sel in menu_selectors:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=1_500):
                    btn.click()
                    page.wait_for_timeout(600)
                    menu_opened = True
                    break
            except Exception:
                continue

        for sel in view_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible(timeout=1_500):
                    el.click()
                    page.wait_for_timeout(1_000)
                    print(f"    Applied DNG view: {view_name}")
                    return
            except Exception:
                continue

        if menu_opened:
            # Close the menu if we couldn't find the view
            page.keyboard.press("Escape")
    except Exception:
        pass  # Non-fatal — continue with whatever the page shows


def is_dng_url(url: str) -> bool:
    return "/rm/" in url or "doors" in url.lower()


def extract_dng_fields(page: Page) -> list[ExtractedField]:
    """Extract all label/value pairs from a DNG requirement page."""
    page_url = page.url

    # Wait for DNG content to render
    try:
        page.wait_for_selector(
            "[class*='attribute'], [class*='property'], [class*='field'], "
            "[class*='label'], [class*='artifact'], table",
            timeout=10_000,
        )
    except Exception:
        pass

    raw: list[dict] = page.evaluate(
        """
        () => {
          const pairs = [];
          const seen = new Set();

          // Strategy 1: label+value sibling pairs in tables (most DNG views)
          document.querySelectorAll('tr').forEach(row => {
            const cells = Array.from(row.querySelectorAll('td, th'));
            if (cells.length >= 2) {
              const label = cells[0].innerText.trim().replace(/\\s+/g,' ').replace(/:$/, '');
              const value = cells[1].innerText.trim().replace(/\\s+/g,' ');
              if (label && label.length < 80 && !seen.has(label)) {
                seen.add(label);
                pairs.push({ name: label, value });
              }
            }
          });

          // Strategy 2: elements with data-attribute / aria-label patterns
          document.querySelectorAll('[class*="attribute"], [class*="property"], [class*="field"]').forEach(el => {
            const labelEl = el.querySelector('[class*="label"], [class*="name"], dt, th');
            const valueEl = el.querySelector('[class*="value"], [class*="content"], dd, td');
            if (labelEl && valueEl) {
              const label = labelEl.innerText.trim().replace(/\\s+/g,' ').replace(/:$/, '');
              const value = valueEl.innerText.trim().replace(/\\s+/g,' ');
              if (label && label.length < 80 && !seen.has(label)) {
                seen.add(label);
                pairs.push({ name: label, value });
              }
            }
          });

          // Strategy 3: definition lists
          document.querySelectorAll('dl').forEach(dl => {
            const dts = Array.from(dl.querySelectorAll('dt'));
            const dds = Array.from(dl.querySelectorAll('dd'));
            dts.forEach((dt, i) => {
              const label = dt.innerText.trim().replace(/\\s+/g,' ').replace(/:$/, '');
              const value = dds[i] ? dds[i].innerText.trim().replace(/\\s+/g,' ') : '';
              if (label && label.length < 80 && !seen.has(label)) {
                seen.add(label);
                pairs.push({ name: label, value });
              }
            });
          });

          return pairs;
        }
        """
    )

    fields: list[ExtractedField] = []
    # Also try text-based extraction using DNG field labels
    known_dng = {normalize_label(l).lower(): l for l in DNG_FIELD_LABELS}
    for r in raw:
        name = clean_text(r.get("name", ""))
        value = clean_text(r.get("value", ""))
        if not name:
            continue
        # Prefer the canonical DNG label name if it matches
        canonical = known_dng.get(normalize_label(name).lower(), name)
        fields.append(ExtractedField(page_url=page_url, name=canonical, value=value, source="dng_page"))

    # Whitelist: only keep dng_page fields whose name is in DNG_FIELD_LABELS
    known_dng_lower = {normalize_label(l).lower() for l in DNG_FIELD_LABELS}
    fields = [
        f for f in fields
        if normalize_label(f.name).lower() in known_dng_lower
    ]

    # Supplement with text-based extraction for known DNG labels
    text_fields = extract_known_fields_from_text(page, DNG_FIELD_LABELS)
    existing_names = {f.name for f in fields if f.value}
    for tf in text_fields:
        if tf.name not in existing_names:
            fields.append(tf)

    return merge_fields(fields)


def extract_implements_requirements(page: Page, resolved_swe1_url: str) -> list[ImplementsRequirementItem]:
    """From a SWE1 work item's Links Tab, find Implements Requirement links and extract their fields."""
    base_urls = build_linked_item_tab_urls(resolved_swe1_url)
    links_tab_url = next((u for u in base_urls if "tab.links" in u), None) if base_urls else None

    if links_tab_url:
        goto_and_settle(page, links_tab_url)
        # Wait for the links section to render in EWM's SPA
        try:
            page.wait_for_selector("a[href]", timeout=10_000)
        except Exception:
            pass

    raw_links = extract_links_tab_entries(page)
    # Deduplicate by href before filtering — same link can appear under
    # multiple section labels in the DOM scan
    seen_hrefs: set[str] = set()
    unique_raw: list[WorkItemLink] = []
    for lnk in raw_links:
        if lnk.href not in seen_hrefs:
            seen_hrefs.add(lnk.href)
            unique_raw.append(lnk)
    impl_links = [lnk for lnk in unique_raw if is_implements_requirement_link(lnk)]
    print(f"  Found {len(impl_links)} 'Implements Requirement' link(s) in SWE1 item.")

    items: list[ImplementsRequirementItem] = []
    seen: set[str] = set()
    for lnk in impl_links:
        href = lnk.href
        if href in seen:
            continue
        seen.add(href)
        try:
            resolved = resolve_view_url(page, href)
            all_fields: list[ExtractedField] = []
            if is_dng_url(resolved):
                print(f"    Extracting DNG requirement: {lnk.text}")
                select_dng_view(page)
                all_fields = extract_dng_fields(page)
            else:
                main_urls = [u for u in build_linked_item_tab_urls(resolved) if "tab.links" not in u] or [resolved]
                for url in main_urls:
                    if url != resolved:
                        goto_and_settle(page, url)
                    all_fields.extend(extract_known_fields_from_text(page, WORK_ITEM_FIELD_LABELS))
            items.append(
                ImplementsRequirementItem(
                    link_text=lnk.text,
                    link_href=href,
                    fields=merge_fields(all_fields),
                )
            )
        except Exception as exc:
            print(f"  Warning: could not extract data from Implements Requirement link {href}: {exc}")
    return items


def extract_swe1_items(page: Page, swe1_links: list[WorkItemLink]) -> list[LinkedWorkItem]:
    """For each SWE1 link go straight to its Links Tab and collect Implements Requirement items."""
    items: list[LinkedWorkItem] = []
    seen: set[str] = set()

    for link in swe1_links:
        href = link.href
        if href in seen:
            continue
        seen.add(href)

        try:
            print(f"Extracting SWE1 item: {link.text}")
            # Resolve the itemOid/redirect URL, then jump straight to the Links Tab
            resolved_url = resolve_view_url(page, href)
            impl_reqs = extract_implements_requirements(page, resolved_url)

            items.append(
                LinkedWorkItem(
                    source_link_text=link.text,
                    source_link_href=href,
                    fields=[],  # SWE1 item fields not needed
                    implements_requirements=impl_reqs,
                )
            )
        except Exception as exc:
            print(f"Warning: could not extract data from SWE1 link {href}: {exc}")

    return items


def discover_links(page: Page, root_url: str, limit: int) -> list[str]:
    links = page.locator("a[href]").evaluate_all(
        """
        anchors => anchors
          .map(a => a.href)
          .filter(Boolean)
        """
    )

    seen: set[str] = set()
    internal_links: list[str] = []
    for link in links:
        absolute = urljoin(page.url, link).split("#")[0]
        if absolute in seen or not same_domain(absolute, root_url):
            continue
        seen.add(absolute)
        internal_links.append(absolute)
        if len(internal_links) >= limit:
            break
    return internal_links


def run_agent(config: AgentConfig) -> AgentResult:
    with sync_playwright() as p:
        launch_options = {"headless": not (config.show_browser or config.manual_login)}
        if config.browser == "msedge":
            launch_options["channel"] = "msedge"
        browser = p.chromium.launch(**launch_options)
        context = browser.new_context()
        page = context.new_page()

        if config.manual_login:
            wait_for_manual_login(page, config.login_url)
        else:
            try:
                login(page, config.login_url, config.username, config.password)
            except RuntimeError as exc:
                if "Could not find visible username/password fields" not in str(exc):
                    raise
                print("")
                print("Automatic login fields were not found. Switching to manual login.")
                wait_for_manual_login(page, config.login_url)

        start_url = config.start_url or page.url
        page.goto(start_url, wait_until="load")
        page.wait_for_timeout(800)

        urls = build_work_item_tab_urls(page.url) or [page.url]

        if config.pages > 1:
            urls.extend(discover_links(page, config.login_url, config.pages - 1))

        all_params: list[PageParameter] = []
        all_fields: list[ExtractedField] = []
        all_links: list[WorkItemLink] = []
        visited: set[str] = set()
        for url in urls:
            if url in visited:
                continue
            visited.add(url)
            goto_and_settle(page, url)
            page_params = extract_parameters(page)
            all_params.extend(page_params)
            all_fields.extend(extract_known_fields_from_text(page, WORK_ITEM_FIELD_LABELS))
            all_fields.extend(fields_from_parameters(page_params))
            all_links.extend(
                lnk for lnk in extract_links_tab_entries(page)
                if not is_implements_requirement_link(lnk)
            )

        swe1_links = [link for link in all_links if is_swe1_link(link, config.swe1_keyword)]
        swe1_items = extract_swe1_items(page, swe1_links) if swe1_links else []

        context.close()
        browser.close()
        return AgentResult(fields=merge_fields(all_fields), parameters=all_params, links=all_links, swe1_items=swe1_items)


def write_excel_report(result: AgentResult, excel_path: Path, generated_at: str) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Extracted Fields ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Extracted Fields"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    missing_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    ok_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")

    ws1.append(["Field", "Value", "Status", "Source"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    extracted_map = {f.name: f for f in result.fields}
    for label in WORK_ITEM_FIELD_LABELS:
        extracted = extracted_map.get(label)
        value = extracted.value if extracted and extracted.value else ""
        source = extracted.source if extracted else ""
        status = "OK" if value else "MISSING"
        row = ws1.max_row + 1
        ws1.append([label, value, status, source])
        status_cell = ws1.cell(row=row, column=3)
        status_cell.fill = ok_fill if status == "OK" else missing_fill
        status_cell.font = Font(bold=True)

    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = max(len(str(cell.value or "")) for cell in col) + 4

    # ── Sheet 2: Missing Mandatory Fields ─────────────────────────────────────
    ws2 = wb.create_sheet("Missing Mandatory Fields")
    ws2.append(["Missing Field"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = PatternFill(fill_type="solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center")

    missing = [label for label in WORK_ITEM_FIELD_LABELS if not (extracted_map.get(label) and extracted_map[label].value)]
    for label in missing:
        ws2.append([label])
    ws2.column_dimensions["A"].width = 30

    # Sheet 3: Links Tab
    ws3 = wb.create_sheet("Links Tab")
    ws3.append(["Link Text", "URL", "Section", "Context", "Page"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if result.links:
        for link in sorted(result.links, key=lambda item: (item.section.lower(), item.context.lower(), item.text.lower(), item.href.lower())):
            ws3.append([link.text, link.href, link.section, link.context, link.page_url])
    else:
        ws3.append(["No links found", "", "", "", ""])

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4,
            80,
        )

    # ── Sheet 4: SWE1 Linked Items ─────────────────────────────────────────────
    ws4 = wb.create_sheet("SWE1 Linked Items")
    ws4.append(["Source Link Text", "Source URL", "Field", "Value", "Source"])
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    if result.swe1_items:
        for item in result.swe1_items:
            if item.fields:
                for extracted in item.fields:
                    ws4.append([item.source_link_text, item.source_link_href, extracted.name, extracted.value, extracted.source])
            else:
                ws4.append([item.source_link_text, item.source_link_href, "No fields found", "", ""])
    else:
        ws4.append(["No SWE1 links found", "", "", "", ""])

    for col in ws4.columns:
        ws4.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4,
            80,
        )

    # ── Sheet 5: Implements Requirement Items (pivoted: req=row, field=col) ───
    ws5 = wb.create_sheet("Implements Requirement")
    impl_green = PatternFill(fill_type="solid", fgColor="375623")

    # Collect all requirement items and build ordered field column list
    impl_all_reqs: list[tuple[str, ImplementsRequirementItem]] = []
    if result.swe1_items:
        for item in result.swe1_items:
            for req in item.implements_requirements:
                impl_all_reqs.append((item.source_link_text, req))

    if impl_all_reqs:
        # Build ordered column set from DNG_FIELD_LABELS first, then any extras
        impl_field_set: set[str] = set()
        impl_field_order: list[str] = []
        for _, req in impl_all_reqs:
            for f in req.fields:
                if f.name not in impl_field_set:
                    impl_field_set.add(f.name)
                    impl_field_order.append(f.name)
        ordered_impl = [f for f in DNG_FIELD_LABELS if f in impl_field_set]
        ordered_impl += [f for f in impl_field_order if f not in ordered_impl]

        header_row5 = ["SWE1 Source", "Req Link Text", "Req URL"] + ordered_impl
        ws5.append(header_row5)
        for cell in ws5[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = impl_green
            cell.alignment = Alignment(horizontal="center")

        for swe1_text, req in impl_all_reqs:
            field_map = {f.name: f.value for f in req.fields}
            row = [swe1_text, req.link_text, req.link_href] + [field_map.get(col, "") for col in ordered_impl]
            ws5.append(row)
    else:
        ws5.append(["No Implements Requirement links found"])

    for col in ws5.columns:
        ws5.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4,
            80,
        )

    # ── Sheet 6: DNG Requirements (MPCI_standard view) ────────────────────────
    # One row per DNG requirement, one column per field — easy to read and compare.
    ws6 = wb.create_sheet("DNG Requirements")
    dng_green = PatternFill(fill_type="solid", fgColor="1F4E79")

    # Collect all DNG requirements across all SWE1 items
    dng_reqs: list[ImplementsRequirementItem] = []
    if result.swe1_items:
        for item in result.swe1_items:
            for req in item.implements_requirements:
                dng_reqs.append(req)

    if dng_reqs:
        # Build ordered column list: key identifier fields first, then rest
        priority_fields = [
            "Artifact Type", "Feature", "SW_Function", "Status",
            "ValidSince", "InvalidSince", "Safety", "Security", "Legal",
            "Verification_Level", "Verification_Criteria",
            "Satisfies", "Link To", "Satisfied by Architecture Element",
            "Implemented By", "Validated By", "RT0ID",
            "Change Request Link", "Resource", "Comment_Internal",
        ]
        all_field_names: list[str] = []
        field_name_set: set[str] = set()
        for req in dng_reqs:
            for f in req.fields:
                if f.name not in field_name_set:
                    field_name_set.add(f.name)
                    all_field_names.append(f.name)
        ordered = [f for f in priority_fields if f in field_name_set]
        ordered += sorted(f for f in all_field_names if f not in ordered)

        header_row = ["Req Link Text", "Req URL"] + ordered
        ws6.append(header_row)
        for cell in ws6[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = dng_green
            cell.alignment = Alignment(horizontal="center")

        for req in dng_reqs:
            field_map = {f.name: f.value for f in req.fields}
            row = [req.link_text, req.link_href] + [field_map.get(col, "") for col in ordered]
            ws6.append(row)

        for col in ws6.columns:
            ws6.column_dimensions[col[0].column_letter].width = min(
                max(len(str(cell.value or "")) for cell in col) + 4,
                80,
            )
    else:
        ws6.append(["No DNG requirements extracted yet."])

    wb.save(excel_path)


def write_reports(result: AgentResult, report_path: str) -> tuple[Path, Path]:
    markdown_path = Path(report_path)
    if markdown_path.suffix.lower() != ".md":
        markdown_path = markdown_path.with_suffix(".md")
    json_path = markdown_path.with_suffix(".json")

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# Work Item Field Report",
        "",
        f"Generated: {generated_at}",
        "",
        "## Extracted Fields",
        "",
        "| Field | Value | Source |",
        "| --- | --- | --- |",
    ]

    extracted_names = {f.name for f in result.fields if f.value}
    missing_mandatory = [label for label in WORK_ITEM_FIELD_LABELS if label not in extracted_names]

    if result.fields:
        for extracted in result.fields:
            value = extracted.value or "Unassigned / empty"
            lines.append(f"| {escape_markdown_table(extracted.name)} | {escape_markdown_table(value)} | {extracted.source} |")
    else:
        lines.append("| No fields found |  |  |")

    if missing_mandatory:
        lines.extend(
            [
                "",
                "## Missing Mandatory Fields",
                "",
                "> The following mandatory fields were **not found** or had **no value** on the inspected page:",
                "",
            ]
        )
        for label in missing_mandatory:
            lines.append(f"- {label}")

    lines.extend(
        [
            "",
            "## Links Tab",
            "",
            "| Link Text | URL | Section | Context |",
            "| --- | --- | --- | --- |",
        ]
    )

    if result.links:
        for link in sorted(result.links, key=lambda item: (item.section.lower(), item.context.lower(), item.text.lower(), item.href.lower())):
            lines.append(
                "| "
                + " | ".join(
                    [
                        escape_markdown_table(link.text),
                        escape_markdown_table(link.href),
                        escape_markdown_table(link.section),
                        escape_markdown_table(link.context),
                    ]
                )
                + " |"
            )
    else:
        lines.append("| No links found |  |  |")

    lines.extend(
        [
            "",
            "## SWE1 Linked Items",
            "",
            "| Source Link | Field | Value | Source |",
            "| --- | --- | --- | --- |",
        ]
    )

    if result.swe1_items:
        for item in result.swe1_items:
            link_cell = f"[{escape_markdown_table(item.source_link_text)}]({item.source_link_href})"
            if item.fields:
                for extracted in item.fields:
                    lines.append(
                        f"| {link_cell} | {escape_markdown_table(extracted.name)} | {escape_markdown_table(extracted.value or 'Unassigned / empty')} | {extracted.source} |"
                    )
            else:
                lines.append(f"| {link_cell} | No fields found |  |  |")
    else:
        lines.append("| No SWE1 links found |  |  |  |")

    lines.extend(
        [
            "",
            "## Implements Requirement",
            "",
            "| SWE1 Source | Req Link | Field | Value | Source |",
            "| --- | --- | --- | --- | --- |",
        ]
    )

    has_impl_md = False
    if result.swe1_items:
        for item in result.swe1_items:
            for req in item.implements_requirements:
                has_impl_md = True
                req_cell = f"[{escape_markdown_table(req.link_text)}]({req.link_href})"
                swe1_cell = escape_markdown_table(item.source_link_text)
                if req.fields:
                    for extracted in req.fields:
                        lines.append(
                            f"| {swe1_cell} | {req_cell} | {escape_markdown_table(extracted.name)} | {escape_markdown_table(extracted.value or 'Unassigned / empty')} | {extracted.source} |"
                        )
                else:
                    lines.append(f"| {swe1_cell} | {req_cell} | No fields found |  |  |")
    if not has_impl_md:
        lines.append("| No Implements Requirement links found |  |  |  |  |")

    lines.extend(
        [
            "",
            "## Raw Parameters",
            "",
            "| Page | Kind | Name | Label | Type | Value | Options |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
    )

    for parameter in result.parameters:
        lines.append(
            "| "
            + " | ".join(
                [
                    escape_markdown_table(parameter.page_url),
                    escape_markdown_table(parameter.kind),
                    escape_markdown_table(parameter.name),
                    escape_markdown_table(parameter.label),
                    escape_markdown_table(parameter.input_type),
                    escape_markdown_table(parameter.value),
                    escape_markdown_table(", ".join(parameter.options)),
                ]
            )
            + " |"
        )

    excel_path = markdown_path.with_suffix(".xlsx")
    markdown_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    json_path.write_text(json.dumps(asdict(result), indent=2), encoding="utf-8")
    write_excel_report(result, excel_path, generated_at)
    return markdown_path, json_path


def escape_markdown_table(value: str) -> str:
    return clean_text(value).replace("|", "\\|")


def main() -> None:
    load_dotenv()

    parser = argparse.ArgumentParser(description="Log in to an authorized website and print visible page parameters.")
    parser.add_argument("--login-url", help="Corporate login page URL.")
    parser.add_argument("--start-url", help="Page to inspect after login. Defaults to the post-login page.")
    parser.add_argument("--pages", type=int, default=1, help="Number of same-domain pages to inspect.")
    parser.add_argument("--show-browser", action="store_true", help="Run with a visible browser for debugging.")
    parser.add_argument("--manual-login", action="store_true", help="Open the browser and let you complete SSO/login manually.")
    parser.add_argument("--browser", choices=["msedge", "chromium"], default="msedge", help="Browser to launch.")
    parser.add_argument("--report", default="work_item_report.md", help="Markdown report output path.")
    parser.add_argument("--swe1-keyword", default="", help="Keyword to filter SWE1 links (e.g. '1014821: [SWE 1]').")
    parser.add_argument("--no-popup", action="store_true", help="Use CLI/env values instead of showing the popup.")
    args = parser.parse_args()

    if args.no_popup:
        if not args.login_url:
            parser.error("--login-url is required when --no-popup is used.")
        config = AgentConfig(
            login_url=args.login_url,
            start_url=args.start_url,
            username="" if args.manual_login else env("CORP_USERNAME"),
            password="" if args.manual_login else env("CORP_PASSWORD"),
            pages=args.pages,
            show_browser=args.show_browser or args.manual_login,
            manual_login=args.manual_login,
            report_path=args.report,
            browser=args.browser,
            swe1_keyword=args.swe1_keyword,
        )
    else:
        config = prompt_for_config()

    try:
        result = run_agent(config)
    except Exception as exc:
        if not args.no_popup:
            messagebox.showerror("Agent failed", str(exc))
        raise

    markdown_path, json_path = write_reports(result, config.report_path)
    print(json.dumps(asdict(result), indent=2))
    print("")
    excel_path = markdown_path.with_suffix(".xlsx")
    print(f"Markdown report saved to: {markdown_path.resolve()}")
    print(f"JSON report saved to: {json_path.resolve()}")
    print(f"Excel report saved to: {excel_path.resolve()}")


if __name__ == "__main__":
    main()
